from __future__ import annotations

import os
import re
import shutil
import subprocess
import uuid
import zipfile
from pathlib import Path
from typing import Iterable, List, Optional, Set

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import pypdfium2 as pdfium
import pytesseract
import yt_dlp
import fitz
from pdf2docx import Converter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from PIL import Image

APP_DIR = Path(__file__).resolve().parent
TMP_DIR = APP_DIR / "tmp"
STATIC_DIR = APP_DIR / "static"
DEFAULT_TESSERACT_EXE = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
DEFAULT_TESSDATA_DIR = Path(r"C:\Users\karim\AppData\Local\Tesseract-OCR\tessdata")
MAX_FILES = 50
MAX_SIZE_BYTES = 150 * 1024 * 1024

TMP_DIR.mkdir(parents=True, exist_ok=True)
STATIC_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="PDF Nova")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


def _new_job_dir() -> Path:
    job_dir = TMP_DIR / f"job_{uuid.uuid4().hex}"
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def _safe_name(name: str, fallback: str = "file") -> str:
    base = Path(name).stem if name else fallback
    cleaned = re.sub(r"[^a-zA-Z0-9_.-]+", "_", base).strip("._")
    return cleaned or fallback


async def _save_upload(upload: UploadFile, out_dir: Path) -> Path:
    raw_name = upload.filename or f"upload_{uuid.uuid4().hex}.bin"
    output = out_dir / f"{_safe_name(raw_name)}{Path(raw_name).suffix.lower()}"
    content = await upload.read()
    if len(content) > MAX_SIZE_BYTES:
        raise HTTPException(status_code=413, detail=f"Fichier trop grand: {raw_name}")
    output.write_bytes(content)
    return output


def _cleanup(path: Path) -> None:
    try:
        if path.is_dir():
            shutil.rmtree(path, ignore_errors=True)
        elif path.exists():
            path.unlink(missing_ok=True)
    except Exception:
        pass


def _file_response(path: Path, filename: str, cleanup_path: Path | None = None) -> FileResponse:
    if cleanup_path is not None:
        # FastAPI does not expose a standard after-send hook on FileResponse.
        # Cleanup is best effort at startup and by replacing old jobs manually.
        pass
    return FileResponse(path=str(path), filename=filename, media_type="application/octet-stream")


def _parse_page_spec(spec: str, total_pages: int) -> List[int]:
    result: Set[int] = set()
    chunks = [p.strip() for p in spec.split(",") if p.strip()]
    if not chunks:
        raise HTTPException(status_code=400, detail="Spec de pages vide.")
    for chunk in chunks:
        if "-" in chunk:
            a_str, b_str = [x.strip() for x in chunk.split("-", 1)]
            if not (a_str.isdigit() and b_str.isdigit()):
                raise HTTPException(status_code=400, detail=f"Intervalle invalide: {chunk}")
            a, b = int(a_str), int(b_str)
            if a > b:
                a, b = b, a
            for p in range(a, b + 1):
                if p < 1 or p > total_pages:
                    raise HTTPException(status_code=400, detail=f"Page hors limite: {p}")
                result.add(p - 1)
        else:
            if not chunk.isdigit():
                raise HTTPException(status_code=400, detail=f"Page invalide: {chunk}")
            p = int(chunk)
            if p < 1 or p > total_pages:
                raise HTTPException(status_code=400, detail=f"Page hors limite: {p}")
            result.add(p - 1)
    return sorted(result)


def _iter_images(paths: Iterable[Path]) -> List[Image.Image]:
    images: List[Image.Image] = []
    for path in paths:
        try:
            img = Image.open(path).convert("RGB")
            images.append(img)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Image invalide: {path.name}") from exc
    if not images:
        raise HTTPException(status_code=400, detail="Aucune image valide.")
    return images


def _page_content_size(page) -> int:
    try:
        contents = page.get_contents()
        if contents is None:
            return 0
        if isinstance(contents, list):
            total = 0
            for item in contents:
                try:
                    total += len(item.get_data())
                except Exception:
                    pass
            return total
        return len(contents.get_data())
    except Exception:
        return 0


def _is_blank_page(page, content_threshold: int) -> bool:
    try:
        text = (page.extract_text() or "").strip()
    except Exception:
        text = ""
    has_annots = page.get("/Annots") is not None
    try:
        has_images = len(list(page.images)) > 0
    except Exception:
        has_images = False
    content_size = _page_content_size(page)
    return (not text) and (not has_annots) and (not has_images) and content_size <= content_threshold


def _ensure_tesseract_available() -> None:
    _configure_tesseract_runtime()
    try:
        pytesseract.get_tesseract_version()
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tesseract non detecte. Installe Tesseract OCR puis ajoute-le au PATH Windows."
            ),
        ) from exc


def _configure_tesseract_runtime() -> None:
    if DEFAULT_TESSERACT_EXE.exists():
        pytesseract.pytesseract.tesseract_cmd = str(DEFAULT_TESSERACT_EXE)
    if DEFAULT_TESSDATA_DIR.exists():
        os.environ.setdefault("TESSDATA_PREFIX", str(DEFAULT_TESSDATA_DIR))


def _ocr_pdf_page_image(src_pdf: Path, index: int) -> Image.Image:
    try:
        doc = pdfium.PdfDocument(str(src_pdf))
        page = doc[index]
        pil = page.render(scale=2.2).to_pil()
        page.close()
        doc.close()
        return pil
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Impossible de rasteriser la page {index + 1}.") from exc


def _is_ffmpeg_available() -> bool:
    try:
        proc = subprocess.run(
            ["ffmpeg", "-version"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        return proc.returncode == 0
    except Exception:
        return False


def _find_soffice() -> Optional[str]:
    candidates = [
        "soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for cand in candidates:
        if cand == "soffice":
            try:
                proc = subprocess.run(
                    [cand, "--version"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    check=False,
                )
                if proc.returncode == 0:
                    return cand
            except Exception:
                continue
        else:
            if Path(cand).exists():
                return cand
    return None


def _prepare_excel_single_page(src: Path, out_dir: Path) -> Path:
    """Force workbook print settings to fit each sheet on one page."""
    out = out_dir / "single_page_input.xlsx"
    try:
        wb = load_workbook(filename=str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Fichier Excel invalide.") from exc

    for ws in wb.worksheets:
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(str(out))
    return out


def _cluster_positions(values: List[float], tolerance: float) -> List[float]:
    if not values:
        return []
    values = sorted(values)
    clusters: List[List[float]] = [[values[0]]]
    for v in values[1:]:
        if abs(v - clusters[-1][-1]) <= tolerance:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return [sum(c) / len(c) for c in clusters]


def _nearest_index(value: float, centers: List[float]) -> int:
    if not centers:
        return 0
    best_idx = 0
    best_dist = abs(value - centers[0])
    for i in range(1, len(centers)):
        d = abs(value - centers[i])
        if d < best_dist:
            best_dist = d
            best_idx = i
    return best_idx


def _bbox_iou(a: tuple[float, float, float, float], b: tuple[float, float, float, float]) -> float:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    ix0, iy0 = max(ax0, bx0), max(ay0, by0)
    ix1, iy1 = min(ax1, bx1), min(ay1, by1)
    iw, ih = max(0.0, ix1 - ix0), max(0.0, iy1 - iy0)
    inter = iw * ih
    if inter <= 0:
        return 0.0
    area_a = max(0.0, (ax1 - ax0) * (ay1 - ay0))
    area_b = max(0.0, (bx1 - bx0) * (by1 - by0))
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


def _extract_tables_with_fallback(page) -> List:
    """Try multiple strategies to catch ruled and borderless tables."""
    configs = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {"vertical_strategy": "text", "horizontal_strategy": "text", "min_words_vertical": 2, "text_tolerance": 4},
    ]
    found = []
    seen_bboxes: List[tuple[float, float, float, float]] = []
    for cfg in configs:
        try:
            tf = page.find_tables(**cfg)
            tables = list(tf.tables) if tf else []
        except Exception:
            tables = []
        for t in tables:
            bbox = tuple(float(v) for v in t.bbox)
            duplicate = any(_bbox_iou(bbox, sb) > 0.75 for sb in seen_bboxes)
            if duplicate:
                continue
            seen_bboxes.append(bbox)
            found.append(t)
    return found


def _table_shape_quality(data: List[List[str]]) -> tuple[int, int, int]:
    rows = len(data or [])
    cols = max((len(r) for r in (data or [])), default=0)
    non_empty = 0
    for r in data or []:
        for c in r:
            if c is not None and str(c).strip() != "":
                non_empty += 1
    return rows, cols, non_empty


def _merge_fragmented_words(words: List[dict]) -> List[dict]:
    if not words:
        return words
    merged: List[dict] = [dict(words[0])]
    for w in words[1:]:
        prev = merged[-1]
        gap = float(w["x0"]) - float(prev["x1"])
        prev_txt = str(prev["text"])
        cur_txt = str(w["text"])
        should_merge = gap <= 12 and (
            len(prev_txt) <= 3
            or len(cur_txt) <= 3
            or (prev_txt.isalpha() and cur_txt.isalpha())
        )
        if should_merge:
            joiner = "" if (prev_txt.isalpha() and cur_txt.isalpha()) else " "
            prev["text"] = f"{prev_txt}{joiner}{cur_txt}".strip()
            prev["x1"] = float(w["x1"])
        else:
            merged.append(dict(w))
    return merged


def _apply_visible_borders(ws) -> None:
    """Apply clear borders on the used rectangular data area."""
    min_row = None
    max_row = 0
    min_col = None
    max_col = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                continue
            min_row = r if min_row is None else min(min_row, r)
            min_col = c if min_col is None else min(min_col, c)
            max_row = max(max_row, r)
            max_col = max(max_col, c)

    if min_row is None or min_col is None:
        return

    edge = Side(border_style="thin", color="6B7280")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(left=edge, right=edge, top=edge, bottom=edge)


@app.get("/")
def root() -> FileResponse:
    return FileResponse(path=str(STATIC_DIR / "index.html"))


@app.get("/api/health")
def health() -> JSONResponse:
    return JSONResponse({"ok": True, "app": "pdf_nova"})


@app.get("/api/capabilities")
def capabilities() -> JSONResponse:
    _configure_tesseract_runtime()
    try:
        version = str(pytesseract.get_tesseract_version())
        ocr_available = True
        ocr_note = f"Tesseract detecte: {version}"
    except Exception:
        ocr_available = False
        ocr_note = "Tesseract non installe ou absent du PATH."
    return JSONResponse(
        {
            "ocr_available": ocr_available,
            "ocr_note": ocr_note,
            "video_extract_available": True,
            "video_extract_note": "yt-dlp actif. ffmpeg requis pour une fusion optimale audio+video.",
            "office_to_pdf_available": _find_soffice() is not None,
            "office_to_pdf_note": "LibreOffice requis pour DOCX/XLSX/PPTX -> PDF.",
        }
    )


@app.post("/api/merge")
async def merge_pdfs(files: List[UploadFile] = File(...)) -> FileResponse:
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="Ajoute au moins 2 PDFs.")
    if len(files) > MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_FILES} fichiers.")

    job_dir = _new_job_dir()
    writer = PdfWriter()
    try:
        for file in files:
            path = await _save_upload(file, job_dir)
            try:
                reader = PdfReader(str(path))
                for page in reader.pages:
                    writer.add_page(page)
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"PDF invalide: {file.filename}") from exc

        out = job_dir / "merged.pdf"
        with out.open("wb") as fh:
            writer.write(fh)
        return _file_response(out, "pdf_nova_merged.pdf")
    finally:
        pass


@app.post("/api/split")
async def split_pdf(
    file: UploadFile = File(...),
    chunk_size: int = Form(1),
) -> FileResponse:
    if chunk_size < 1:
        raise HTTPException(status_code=400, detail="chunk_size doit etre >= 1")

    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    try:
        reader = PdfReader(str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF invalide.") from exc

    total = len(reader.pages)
    zip_path = job_dir / "split.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        part_idx = 1
        for start in range(0, total, chunk_size):
            writer = PdfWriter()
            end = min(start + chunk_size, total)
            for i in range(start, end):
                writer.add_page(reader.pages[i])
            part_name = f"part_{part_idx:03d}.pdf"
            tmp_pdf = job_dir / part_name
            with tmp_pdf.open("wb") as fh:
                writer.write(fh)
            zf.write(tmp_pdf, part_name)
            part_idx += 1
    return _file_response(zip_path, "pdf_nova_split.zip")


@app.post("/api/extract")
async def extract_pages(
    file: UploadFile = File(...),
    pages: str = Form(...),
) -> FileResponse:
    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    try:
        reader = PdfReader(str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF invalide.") from exc

    indices = _parse_page_spec(pages, len(reader.pages))
    writer = PdfWriter()
    for i in indices:
        writer.add_page(reader.pages[i])
    out = job_dir / "extracted.pdf"
    with out.open("wb") as fh:
        writer.write(fh)
    return _file_response(out, "pdf_nova_extract.pdf")


@app.post("/api/rotate")
async def rotate_pdf(
    file: UploadFile = File(...),
    angle: int = Form(...),
    pages: str = Form(""),
) -> FileResponse:
    if angle not in {90, 180, 270}:
        raise HTTPException(status_code=400, detail="angle doit etre 90, 180 ou 270.")
    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    try:
        reader = PdfReader(str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF invalide.") from exc

    targets = (
        set(_parse_page_spec(pages, len(reader.pages)))
        if pages.strip()
        else set(range(len(reader.pages)))
    )
    writer = PdfWriter()
    for idx, page in enumerate(reader.pages):
        if idx in targets:
            page.rotate(angle)
        writer.add_page(page)
    out = job_dir / "rotated.pdf"
    with out.open("wb") as fh:
        writer.write(fh)
    return _file_response(out, "pdf_nova_rotate.pdf")


@app.post("/api/watermark")
async def watermark_pdf(
    file: UploadFile = File(...),
    text: str = Form(...),
    opacity: float = Form(0.15),
) -> FileResponse:
    text = text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="Texte obligatoire.")
    if opacity <= 0 or opacity > 1:
        raise HTTPException(status_code=400, detail="opacity doit etre entre 0 et 1.")

    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    try:
        reader = PdfReader(str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF invalide.") from exc

    writer = PdfWriter()
    for i, page in enumerate(reader.pages):
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        wm_path = job_dir / f"wm_{i}.pdf"
        c = canvas.Canvas(str(wm_path), pagesize=(width, height))
        c.setFillColor(colors.Color(0.15, 0.15, 0.15, alpha=opacity))
        c.translate(width / 2, height / 2)
        c.rotate(35)
        c.setFont("Helvetica-Bold", min(64, max(24, int(width / 10))))
        c.drawCentredString(0, 0, text)
        c.save()

        wm_reader = PdfReader(str(wm_path))
        wm_page = wm_reader.pages[0]
        page.merge_page(wm_page)
        writer.add_page(page)

    out = job_dir / "watermarked.pdf"
    with out.open("wb") as fh:
        writer.write(fh)
    return _file_response(out, "pdf_nova_watermark.pdf")


@app.post("/api/images-to-pdf")
async def images_to_pdf(files: List[UploadFile] = File(...)) -> FileResponse:
    if not files:
        raise HTTPException(status_code=400, detail="Ajoute au moins 1 image.")
    if len(files) > MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_FILES} fichiers.")

    job_dir = _new_job_dir()
    saved: List[Path] = []
    for file in files:
        saved.append(await _save_upload(file, job_dir))

    images = _iter_images(saved)
    out = job_dir / "images.pdf"
    first, rest = images[0], images[1:]
    first.save(out, "PDF", resolution=100.0, save_all=True, append_images=rest)
    return _file_response(out, "pdf_nova_images.pdf")


@app.post("/api/compress")
async def compress_pdf(
    file: UploadFile = File(...),
    level: str = Form("balanced"),
) -> FileResponse:
    level = level.strip().lower()
    if level not in {"light", "balanced", "aggressive"}:
        raise HTTPException(status_code=400, detail="level invalide: light, balanced, aggressive.")

    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    try:
        reader = PdfReader(str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF invalide.") from exc

    writer = PdfWriter()
    for page in reader.pages:
        try:
            page.compress_content_streams()
        except Exception:
            pass
        writer.add_page(page)

    try:
        writer.compress_identical_objects(remove_identicals=True, remove_orphans=True)
    except Exception:
        pass

    if reader.metadata:
        writer.add_metadata(dict(reader.metadata))

    out = job_dir / "compressed.pdf"
    with out.open("wb") as fh:
        writer.write(fh)
    return _file_response(out, "pdf_nova_compress.pdf")


@app.post("/api/remove-blank")
async def remove_blank_pages(
    file: UploadFile = File(...),
    content_threshold: int = Form(80),
) -> FileResponse:
    if content_threshold < 0:
        raise HTTPException(status_code=400, detail="content_threshold doit etre >= 0.")
    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    try:
        reader = PdfReader(str(src))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="PDF invalide.") from exc

    writer = PdfWriter()
    kept = 0
    for page in reader.pages:
        if not _is_blank_page(page, content_threshold):
            writer.add_page(page)
            kept += 1

    if kept == 0:
        raise HTTPException(status_code=400, detail="Toutes les pages ont ete detectees comme blanches.")

    out = job_dir / "no_blank.pdf"
    with out.open("wb") as fh:
        writer.write(fh)
    return _file_response(out, "pdf_nova_no_blank.pdf")


@app.post("/api/ocr-text")
async def ocr_text(
    file: UploadFile = File(...),
    lang: str = Form("fra+eng"),
    min_chars: int = Form(40),
) -> FileResponse:
    if min_chars < 0:
        raise HTTPException(status_code=400, detail="min_chars doit etre >= 0.")
    _ensure_tesseract_available()

    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    ext = src.suffix.lower()
    blocks: List[str] = []

    if ext == ".pdf":
        try:
            reader = PdfReader(str(src))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="PDF invalide.") from exc
        for idx, page in enumerate(reader.pages):
            extracted = (page.extract_text() or "").strip()
            if len(extracted) >= min_chars:
                text = extracted
            else:
                img = _ocr_pdf_page_image(src, idx)
                text = pytesseract.image_to_string(img, lang=lang).strip()
            blocks.append(f"===== PAGE {idx + 1} =====\n{text}\n")
    else:
        try:
            img = Image.open(src).convert("RGB")
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Format image non supporte.") from exc
        text = pytesseract.image_to_string(img, lang=lang).strip()
        blocks.append(f"===== IMAGE =====\n{text}\n")

    out = job_dir / "ocr.txt"
    out.write_text("\n".join(blocks), encoding="utf-8")
    return _file_response(out, "pdf_nova_ocr.txt")


@app.post("/api/video-extract")
async def video_extract_mp4(
    video_url: str = Form(...),
    source: str = Form(""),
    owns_rights: str = Form("false"),
) -> FileResponse:
    video_url = video_url.strip()
    if not video_url:
        raise HTTPException(status_code=400, detail="URL obligatoire.")
    if owns_rights.strip().lower() not in {"true", "1", "yes", "on"}:
        raise HTTPException(
            status_code=400,
            detail="Confirme que tu as les droits de telechargement.",
        )

    job_dir = _new_job_dir()
    out_tmpl = str(job_dir / "video.%(ext)s")
    ffmpeg_ok = _is_ffmpeg_available()
    ydl_opts = {
        "outtmpl": out_tmpl,
        "noplaylist": True,
        "quiet": True,
        "no_warnings": True,
        "format": "bv*[ext=mp4]+ba[ext=m4a]/b[ext=mp4]/best" if ffmpeg_ok else "best[ext=mp4]/best",
        "merge_output_format": "mp4",
        "restrictfilenames": True,
    }

    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=True)
            path_str = ydl.prepare_filename(info)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Extraction impossible pour ce lien ({source or 'platform'}).",
        ) from exc

    result_path = Path(path_str)
    if result_path.suffix.lower() != ".mp4":
        mp4_path = result_path.with_suffix(".mp4")
        if mp4_path.exists():
            result_path = mp4_path

    if not result_path.exists():
        raise HTTPException(status_code=500, detail="Fichier video non genere.")

    out_name = f"pdf_nova_video_{_safe_name(source or 'social')}.mp4"
    return _file_response(result_path, out_name)


@app.post("/api/convert")
async def convert_file(
    file: UploadFile = File(...),
    mode: str = Form(...),
) -> FileResponse:
    mode = mode.strip().lower()
    supported = {"pdf_to_docx", "pdf_to_images", "pdf_to_excel", "office_to_pdf", "image_to_pdf"}
    if mode not in supported:
        raise HTTPException(status_code=400, detail=f"mode invalide. Utilise: {', '.join(sorted(supported))}")

    job_dir = _new_job_dir()
    src = await _save_upload(file, job_dir)
    ext = src.suffix.lower()

    if mode == "pdf_to_docx":
        if ext != ".pdf":
            raise HTTPException(status_code=400, detail="Mode pdf_to_docx: fichier PDF requis.")
        out = job_dir / "converted.docx"
        try:
            cv = Converter(str(src))
            cv.convert(str(out))
            cv.close()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Conversion PDF -> DOCX impossible pour ce fichier.") from exc
        return _file_response(out, "pdf_nova_converted.docx")

    if mode == "pdf_to_images":
        if ext != ".pdf":
            raise HTTPException(status_code=400, detail="Mode pdf_to_images: fichier PDF requis.")
        try:
            doc = pdfium.PdfDocument(str(src))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="PDF invalide.") from exc
        zip_path = job_dir / "pdf_images.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for i in range(len(doc)):
                page = doc[i]
                img = page.render(scale=2.0).to_pil()
                page_path = job_dir / f"page_{i + 1:03d}.png"
                img.save(page_path, "PNG")
                zf.write(page_path, page_path.name)
                page.close()
        doc.close()
        return _file_response(zip_path, "pdf_nova_pages.zip")

    if mode == "pdf_to_excel":
        if ext != ".pdf":
            raise HTTPException(status_code=400, detail="Mode pdf_to_excel: fichier PDF requis.")
        try:
            reader = PdfReader(str(src))
            fdoc = fitz.open(str(src))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="PDF invalide.") from exc

        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)
        thin = Side(border_style="thin", color="D0D7E2")

        # Better fidelity approach:
        # 1) Detect tables and write true rows/cols with borders.
        # 2) Add non-table text lines with style hints.
        for idx in range(len(fdoc)):
            ws = wb.create_sheet(f"Page_{idx + 1}")
            page = fdoc[idx]
            row_cursor = 1

            raw_tables = _extract_tables_with_fallback(page)
            table_objs = []
            table_bboxes = []
            for t in raw_tables:
                data = t.extract() or []
                rows, cols, non_empty = _table_shape_quality(data)
                # Ignore weak / false-positive table detections.
                if rows < 2 or cols < 2 or non_empty < 4:
                    continue
                table_objs.append({"bbox": tuple(t.bbox), "data": data})
                table_bboxes.append(tuple(t.bbox))

            def _inside_table(x: float, y: float) -> bool:
                for bx0, by0, bx1, by1 in table_bboxes:
                    if (bx0 - 1) <= x <= (bx1 + 1) and (by0 - 1) <= y <= (by1 + 1):
                        return True
                return False

            elements = []
            for t in table_objs:
                elements.append({"kind": "table", "y": float(t["bbox"][1]), "table": t})

            words = page.get_text("words")
            free_words = []
            for w in words:
                x0, y0, x1, y1, txt = float(w[0]), float(w[1]), float(w[2]), float(w[3]), str(w[4]).strip()
                if not txt:
                    continue
                cx, cy = (x0 + x1) / 2, (y0 + y1) / 2
                if _inside_table(cx, cy):
                    continue
                free_words.append({"x0": x0, "x1": x1, "y0": y0, "text": txt})

            if free_words:
                y_centers = _cluster_positions([w["y0"] for w in free_words], tolerance=3.2)
                lines_map: dict[int, list] = {}
                for w in free_words:
                    ridx = _nearest_index(w["y0"], y_centers)
                    lines_map.setdefault(ridx, []).append(w)
                for ridx, line_words in lines_map.items():
                    line_words.sort(key=lambda s: s["x0"])
                    line_words = _merge_fragmented_words(line_words)
                    elements.append({"kind": "line", "y": y_centers[ridx], "spans": line_words})

            elements.sort(key=lambda e: e["y"])

            if not elements:
                ws["A1"] = "No extractable text found on this page."
                continue

            for el in elements:
                if el["kind"] == "line":
                    col = 1
                    spans = el["spans"]
                    for i, sp in enumerate(spans):
                        cell = ws.cell(row=row_cursor, column=col, value=sp["text"])
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
                        col += 1
                        if i < len(spans) - 1:
                            gap = spans[i + 1]["x0"] - sp["x1"]
                            if gap > 80:
                                col += 1
                    row_cursor += 1
                    continue

                data = el["table"]["data"]
                for ridx, row_vals in enumerate(data, start=0):
                    for cidx, raw_val in enumerate(row_vals, start=1):
                        val = raw_val.strip() if isinstance(raw_val, str) else raw_val
                        cell = ws.cell(row=row_cursor + ridx, column=cidx, value=val)
                        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        if ridx == 0:
                            cell.font = Font(bold=True)
                row_cursor += max(1, len(data)) + 1

            # Autofit widths roughly by content length.
            max_col = min(ws.max_column, 60)
            for c in range(1, max_col + 1):
                mx = 0
                for r in range(1, min(ws.max_row, 4000) + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    ln = len(str(v))
                    if ln > mx:
                        mx = ln
                ws.column_dimensions[get_column_letter(c)].width = min(50, max(10, mx * 0.95))

            _apply_visible_borders(ws)

        # Optional flat table for quick filtering/searching.
        summary = wb.create_sheet("All_Text")
        summary.append(["page", "line", "content"])
        for page_idx, page in enumerate(reader.pages, start=1):
            lines = [ln.strip() for ln in (page.extract_text() or "").splitlines() if ln.strip()]
            if not lines:
                summary.append([page_idx, 1, ""])
            else:
                for line_idx, line in enumerate(lines, start=1):
                    summary.append([page_idx, line_idx, line])

        out = job_dir / "converted.xlsx"
        wb.save(out)
        fdoc.close()
        return _file_response(out, "pdf_nova_converted.xlsx")

    if mode == "image_to_pdf":
        if ext not in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff", ".tif"}:
            raise HTTPException(status_code=400, detail="Mode image_to_pdf: fichier image requis.")
        out = job_dir / "image.pdf"
        try:
            img = Image.open(src).convert("RGB")
            img.save(out, "PDF", resolution=100.0)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Conversion Image -> PDF impossible.") from exc
        return _file_response(out, "pdf_nova_image.pdf")

    # office_to_pdf
    if ext not in {".docx", ".xlsx", ".pptx"}:
        raise HTTPException(status_code=400, detail="Mode office_to_pdf: fichier DOCX/XLSX/PPTX requis.")
    soffice = _find_soffice()
    if not soffice:
        raise HTTPException(status_code=400, detail="LibreOffice non detecte. Installe LibreOffice pour convertir en PDF.")

    source_for_lo = src
    if ext == ".xlsx":
        source_for_lo = _prepare_excel_single_page(src, job_dir)

    try:
        proc = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(job_dir), str(source_for_lo)],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
            text=True,
            timeout=180,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Echec execution LibreOffice.") from exc

    if proc.returncode != 0:
        raise HTTPException(status_code=400, detail="Conversion Office -> PDF echouee.")

    out_stem = source_for_lo.stem if ext == ".xlsx" else src.stem
    out = job_dir / f"{out_stem}.pdf"
    if ext == ".xlsx" and not out.exists():
        # Some LibreOffice builds may still output using original filename.
        alt = job_dir / f"{src.stem}.pdf"
        if alt.exists():
            out = alt
    if not out.exists():
        raise HTTPException(status_code=500, detail="PDF converti introuvable.")
    filename = "pdf_nova_excel_single_page.pdf" if ext == ".xlsx" else "pdf_nova_office.pdf"
    return _file_response(out, filename)


@app.on_event("startup")
def cleanup_old_jobs() -> None:
    _configure_tesseract_runtime()
    for item in TMP_DIR.glob("job_*"):
        if item.is_dir():
            try:
                shutil.rmtree(item, ignore_errors=True)
            except Exception:
                pass


if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", "8091"))
    uvicorn.run("server:app", host="0.0.0.0", port=port, reload=False)
